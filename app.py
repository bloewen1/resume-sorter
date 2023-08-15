from ctypes import cdll, c_char_p, c_size_t
from flask import Flask, render_template, request, jsonify, redirect, send_file
from io import BytesIO
from docx import Document
import PyPDF2
import zipfile
import openpyxl
import re
from datetime import datetime
import re
from docx.shared import Pt, RGBColor
import string

# Create a Flask app and links to the static folder
app = Flask(__name__, static_url_path='/static')

# Load the Rust shared library
rust_lib_path = r"C:\Users\aaron.laitner\Source\Repos\resume-sorter\rust_code\rust_keyword_analysis_lib\target\release\rust_keyword_analysis_lib.dll"
rust_lib = cdll.LoadLibrary(rust_lib_path)

# Define the analyze_keywords function signature
rust_lib.analyze_keywords.argtypes = [c_char_p, c_size_t]
rust_lib.analyze_keywords.restype = c_char_p

# Define a list of keywords to search for
keywords = []
# Defines a dictionary of roles and keywords
roles = {
    "General": ["Team", "Diploma"],
    "Developer": ["Python", "Javascript", "SQL", "HTML", "Oracle"],
    "Project Manager": ["Manage"],
    "Other": ["Admin"]
}

# If there's an existing workbook, open it, otherwise make a new one
try:
    wb = openpyxl.load_workbook("resumes.xlsx")
except:
    wb = openpyxl.Workbook()

if "Resumes" in wb.sheetnames:
    ws = wb["Resumes"]
else:
    ws = wb.create_sheet("Resumes")
    ws.append(["Filename", "Keywords"])
if "Roles" in wb.sheetnames:
    ws2 = wb["Roles"]
else:
    ws2 = wb.create_sheet("Roles")
    ws2.append(["Role", "Keywords"])

for row in ws2.iter_rows(min_row=2, values_only=True):
    roles[row[0]].append(row[1])

keywords = roles["General"] + roles["Developer"] + roles["Project Manager"] + roles["Other"]

@app.route('/')
def index():
    return render_template('index.html', keywords=keywords, roles=roles)

@app.route('/analyze_keywords', methods=['POST'])
def analyze_keywords():
    contents = request.data
    contents_len = len(contents)
    result_ptr = rust_lib.analyze_keywords(contents, contents_len)
    result = result_ptr.decode()
    return result

@app.route('/add_keyword', methods=['POST'])
def add_keyword():
    role = request.form.get('role')
    new_keyword = request.form.get('new_keyword')
    if new_keyword:
        if new_keyword not in keywords:
            ws2.append([role, new_keyword])
            wb.save("resumes.xlsx")
    return redirect('/')

@app.route('/delete_row', methods=['POST'])
def delete_row():
    try:
        data = request.get_json()  # Get JSON data sent from client
        filename = data.get('filename')

        for row in ws.iter_rows(min_row=2):
            if row[0].value == filename:  # Find the row with the specified filename
                ws.delete_rows(row[0].row)
                wb.save("resumes.xlsx")
                return jsonify({"success": True})
    except Exception as e:
        print("Error deleting row:", e)

    return jsonify({"success": False})

@app.route('/delete', methods=['POST'])
def delete():
    try:
        ws.delete_rows(2, ws.max_row - 1)  # Delete all rows except the header
        wb.save("resumes.xlsx")
    except Exception as e:
        print("Error deleting data:", e)

    return redirect('/')

@app.route('/score', methods=['POST'])
def rank_files():
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        filename, keywords_str = row
        keys = keywords_str.split(', ') if keywords_str else []
        data.append({"filename": filename, "keywords": keys})
    
    # Get the list of search words from the form
    search_words = request.form.getlist('search_word')

    # Check if search words are provided
    if not search_words or len(search_words) == 0:
        search_words = keywords

    results = {}
    for entry in data:
        filename = entry["filename"]
        keys = entry["keywords"]
        found_words = []
        found_count = 0
        for word in search_words:
            if word.lower() in [kw.lower() for kw in keys]:
                found_words.append(word)
                found_count += 1
        total_words = len(search_words)
        score = round((found_count / total_words) * 100)
        results[filename] = {"score": score, "keywords": found_words, "date": get_date_from_filename(filename)}

     # Sort the results based on the score in descending order
    sorted_results = dict(sorted(results.items(), key=lambda item: (item[1]['score'], item[1]['date']), reverse=True))
    
    # Extract file names and scores separately
    file_names = list(sorted_results.keys())
    scores = [sorted_results[filename]['score'] for filename in file_names]
    keys = [sorted_results[filename]['keywords'] for filename in file_names]

    # Zip the file names and scores together
    file_scores = list(zip(file_names, scores, keys))

    return jsonify(file_scores)

@app.route('/parse', methods=['POST'])
def parse_files():
    # Check if 'file' is present in the request
    if 'file' not in request.files:
        return jsonify({"error": "No file part"})

    # Get the uploaded zip file
    zip_file = request.files['file']

    # Check if the uploaded file is a valid zip file
    if not zipfile.is_zipfile(zip_file):
        return jsonify({"error": "Invalid zip file"})

    try:
        # Extract the contents of the zip file
        file_contents = {}
        with zipfile.ZipFile(zip_file, 'r') as zip_archive:
            for filename in zip_archive.namelist():
                with zip_archive.open(filename) as file:
                    file_contents[filename] = file.read()
    except zipfile.BadZipFile:
        return jsonify({"error": "Bad Zip File"})

    # Process each file in the zip
    for filename, file_content in file_contents.items():
        if filename.endswith('.docx'):
            parsed_content = parse_word_document(file_content)
        elif filename.endswith('.pdf'):
            parsed_content = parse_pdf(file_content)
        else:
            continue

        found_words = []
        for key in keywords:
            if remove_punctuation(key.lower()) in parsed_content.lower():
                found_words.append(key)
        file_keywords = ', '.join(found_words)

        existing_row_index = None
        for row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row[0] == filename:
                existing_row_index = row_index
                break

        if existing_row_index:
            # Replace existing row if filename already exists
            ws.delete_rows(existing_row_index)
            new_row = [filename, file_keywords]
            ws.insert_rows(existing_row_index, amount=1)
            for col_index, value in enumerate(new_row, start=1):
                ws.cell(row=existing_row_index, column=col_index, value=value)
        else:
            # Add new row if filename doesn't exist
            ws.append([filename, file_keywords])

    wb.save("resumes.xlsx")

    return jsonify(True)

@app.route('/download', methods=['GET'])
def download_excel():
    try:
        return send_file("resumes.xlsx", as_attachment=True)
    except Exception as e:
        print("Error downloading Excel file:", e)
        return jsonify({"error": "Failed to download Excel file"})

def parse_word_document(file_content):
    # Parse the content of a Word document
    doc = Document(BytesIO(file_content))
    content = ""
    for paragraph in doc.paragraphs:
        # Reset font attributes to default values
        for run in paragraph.runs:
            run.font.size = Pt(12)  # Reset font size to a default value
            run.font.color.rgb = RGBColor(0, 0, 0)  # Reset font color to black
            
        # Add plain text content of the paragraph
        content += paragraph.text + "\n"
    
    # Remove extra spaces and formatting characters
    cleaned_content = re.sub(r'\s+', ' ', content).strip()
    cleaned_content = remove_punctuation(cleaned_content)
    return cleaned_content

def parse_pdf(file_content):
    # Parse the content of a PDF document
    content = ""
    try:
        pdf_file = PyPDF2.PdfReader(BytesIO(file_content))
        for page_num in range(len(pdf_file.pages)):
            page = pdf_file.pages[page_num]
            content += page.extract_text() + "\n"
    except Exception as e:
        print("Error parsing PDF: ", e)
    return remove_punctuation(content)

def get_date_from_filename(filename):
    if filename is None:
        return datetime.min  # Return a default date if filename is None
    if filename is None:
        return datetime.min  # Return a default date if filename is None
    match = re.search(r'\d{4}-\d{2}-\d{2}', filename)
    if match:
        date_str = match.group(0)
        return datetime.strptime(date_str, '%Y-%m-%d')
    return datetime.min  # Return a default date if no match is found

def remove_punctuation(input_string):
    # Define a string containing all punctuation characters
    punctuation = string.punctuation + "“”‘’"  # Adding unicode smart quotes
    
    # Remove punctuation using string translation
    translator = str.maketrans('', '', punctuation)
    cleaned_string = input_string.translate(translator)
    
    # Alternatively, you can use regular expressions to remove punctuation
    cleaned_string = re.sub(r'[{}]'.format(re.escape(punctuation)), '', cleaned_string)
    
    return cleaned_string

if __name__ == '__main__':
    # Set the maximum content length for file uploads to 250 MB
    app.config['MAX_CONTENT_LENGTH'] = 250 * 1024 * 1024
    app.run(debug=True)